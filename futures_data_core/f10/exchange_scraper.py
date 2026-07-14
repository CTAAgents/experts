"""交易所官方数据爬取助手 [INDEPENDENT]。

为仓单 / 行情等模块提供通用的「HTTP 抓取 + 表格解析」能力。交易所端点与数据格式
取自 futures-data-search 已验证的 ``exchange_api_config.json``（SHFE=json, DCE=tsv,
CZCE/GFEX=html, CFFEX=csv），内嵌为常量以避免对外部文件的不确定依赖。

所有网络访问通过**可注入 transport** 完成；解析函数为纯函数便于单元测试。
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime
from typing import Awaitable, Callable, Optional, Union

try:
    from bs4 import BeautifulSoup

    _HAS_BS4 = True
except ImportError:  # pragma: no cover - bs4 为声明依赖，正常环境必有
    BeautifulSoup = None
    _HAS_BS4 = False


# 端点与格式（源自已验证的 exchange_api_config.json）
EXCHANGE_ENDPOINTS: dict[str, dict] = {
    "SHFE": {
        "base_url": "http://www.shfe.com.cn",
        "fmt": "json",
        "endpoint": "/data/dailydata/kx/kx{date}.dat",
    },
    "DCE": {
        "base_url": "http://www.dce.com.cn",
        "fmt": "tsv",
        "endpoint": "/publicweb/quotesdata/exportDayQuotesChData.html",
    },
    "CZCE": {
        "base_url": "http://www.czce.com.cn",
        "fmt": "html",
        "endpoint": "/cn/DFSStaticFiles/Future/{year}/{date}/FutureDataDaily.htm",
    },
    "CFFEX": {
        "base_url": "http://www.cffex.com.cn",
        "fmt": "csv",
        "endpoint": "/sj/hqsj/rtj/{year_month}/{day}/{date}_1.csv",
    },
    "GFEX": {
        "base_url": "http://www.gfex.com.cn",
        "fmt": "html",
        "endpoint": "/gfex/rihq/{date}.js",
    },
}

# 仓单日报专用端点（独立于日线数据端点）
WARRANT_ENDPOINTS: dict[str, dict] = {
    "SHFE": {
        "base_url": "http://www.shfe.com.cn",
        "fmt": "json",
        "endpoint": "/data/dailydata/stocks/{date}dailystock.dat",
    },
    "DCE": {
        "base_url": "http://www.dce.com.cn",
        "fmt": "tsv",
        "endpoint": "/publicweb/quotesdata/exportDayQuotesChData.html",
    },
    "CZCE": {
        "base_url": "http://www.czce.com.cn",
        "fmt": "xlsx",
        "endpoint": "/cn/DFSStaticFiles/Future/{year}/{date}/FutureDataWhsheet.xlsx",
    },
    "GFEX": {
        "base_url": "http://www.gfex.com.cn",
        "fmt": "html",
        "endpoint": "/gfex/rihq/{date}.js",
    },
}

# 持仓排名专用端点（独立于日线/仓单）。
# URL / 请求方式严格对齐 AKShare 已验证实现（同源于交易所官网）。
# 2026-07-14 实盘校准：此前模板为猜测值（多已 404），以下为真实端点。
POSITION_RANK_ENDPOINTS: dict[str, dict] = {
    # 上期所：JSON GET，需 shfe_headers；返回 {"o_cursor":[{...}]}
    "SHFE": {
        "method": "get",
        "fmt": "json",
        "url_tmpl": "https://www.shfe.com.cn/data/tradedata/future/dailydata/pm{date}.dat",
        "headers": {"User-Agent": "Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)"},
    },
    # 中金所：CSV GET，路径含 品种；列含 合约/排名/买卖持仓
    "CFFEX": {
        "method": "get",
        "fmt": "csv",
        "url_tmpl": "http://www.cffex.com.cn/sj/ccpm/{year_month}/{day}/{VAR}_1.csv",
        "headers": {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
    },
    # 郑商所：xlsx GET（2025-11 后统一 xlsx），多品种块，需 pandas 解析（暂未直连→AKShare降级）
    "CZCE": {
        "method": "get",
        "fmt": "xlsx",
        "url_tmpl": "http://www.czce.com.cn/cn/DFSStaticFiles/Future/{year}/{date}/FutureDataHolding.xlsx",
        "headers": {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
        "needs_bytes": True,
    },
    # 大商所：txt GET 需 品种+合约；多合约需循环（暂未直连→AKShare降级）
    "DCE": {
        "method": "get",
        "fmt": "txt",
        "url_tmpl": "http://portal.dce.com.cn/publicweb/quotesdata/exportMemberDealPosiQuotesData.html"
        "?memberDealPosiQuotes.variety={var}&memberDealPosiQuotes.trade_type=0"
        "&contract.contract_id={contract}&contract.variety_id={var}"
        "&year={year}&month={month0}&day={day}&exportFlag=txt",
        "needs_contract": True,
    },
    # 广期所：JSON POST，多步（品种→合约→数据）（暂未直连→AKShare降级）
    "GFEX": {
        "method": "post",
        "fmt": "json",
        "url_tmpl": "http://www.gfex.com.cn/u/interfacesWebTiMemberDealPosiQuotes/loadList",
        "needs_contract": True,
    },
}


def get_position_rank_url(
    exchange: str, trade_date: str, variety: str = "", contract: Optional[str] = None
) -> Optional[str]:
    """持仓排名页面 URL；未知交易所或缺少必要参数（如 DCE/GFEX 需合约）返回 ``None``。"""
    ex = POSITION_RANK_ENDPOINTS.get(exchange.upper())
    if ex is None:
        return None
    if ex.get("needs_contract") and not contract:
        return None
    y = trade_date[:4]
    ym = trade_date[:6]
    day = trade_date[6:]
    month0 = str(int(trade_date[4:6]) - 1)  # DCE month 为 0-based
    return (
        ex["url_tmpl"]
        .replace("{year}", y)
        .replace("{year_month}", ym)
        .replace("{date}", trade_date)
        .replace("{day}", day)
        .replace("{var}", (variety or "").lower())
        .replace("{VAR}", (variety or "").upper())
        .replace("{contract}", (contract or "").lower())
        .replace("{month0}", month0)
    )


def rank_fmt_of(exchange: str) -> Optional[str]:
    """返回持仓排名数据格式（json/csv/tsv/html/xlsx）。"""
    ex = POSITION_RANK_ENDPOINTS.get(exchange.upper())
    return ex["fmt"] if ex else None


def rank_headers_of(exchange: str) -> Optional[dict]:
    """返回持仓排名请求头（如 SHFE 需特定 UA）。"""
    ex = POSITION_RANK_ENDPOINTS.get(exchange.upper())
    return ex.get("headers") if ex else None


def parse_position_rank(
    text_or_bytes: Union[str, bytes], fmt: str, exchange: str, variety: str
) -> dict:
    """解析交易所持仓排名文本/字节为结构化 dict。

    返回 ``{"long":[{rank,member,lots}], "short":[...], "net_position":int,
    "contract":str}``；无法解析返回空 dict。

    直连实现：SHFE(JSON)、CFFEX(CSV)、CZCE(xlsx)、DCE(txt→excel 逐合约)、GFEX(JSON POST 逐合约)。
    """
    exchange_u = exchange.upper()
    if exchange_u == "SHFE":
        assert isinstance(text_or_bytes, str)
        return _parse_shfe_rank(text_or_bytes, variety)
    if exchange_u == "CFFEX":
        # CFFEX 接受 str 或 bytes（GBK 编码），_parse_cffex_rank 自动处理
        return _parse_cffex_rank(text_or_bytes, variety)
    if exchange_u == "CZCE":
        assert isinstance(text_or_bytes, bytes), "CZCE rank requires bytes (xlsx)"
        return _parse_czce_rank_xlsx(text_or_bytes, variety)
    if exchange_u == "DCE":
        if isinstance(text_or_bytes, str):
            return _parse_dce_rank_text(text_or_bytes, variety)
        return _parse_dce_rank_xlsx(text_or_bytes, variety)
    if exchange_u == "GFEX":
        assert isinstance(text_or_bytes, str)
        return _parse_gfex_rank(text_or_bytes, variety)
    return {}


def _parse_shfe_rank(text: str, variety: str) -> dict:
    """解析 SHFE pm{date}.dat JSON（``o_cursor`` 列表）。

    真实列：RANK / INSTRUMENTID(合约) / PARTICIPANTABBR1(成交量会员)
    CJ1(成交量) CJ1_CHG / PARTICIPANTABBR2(持多单会员) CJ2(持多单) CJ2_CHG
    / PARTICIPANTABBR3(持空单会员) CJ3(持空单) CJ3_CHG
    """
    try:
        data = json.loads(text)
    except (ValueError, TypeError):
        return {}
    if not isinstance(data, dict):
        return {}
    rows = data.get("o_cursor") or []
    if not isinstance(rows, list) or not rows:
        return {}
    var_u = variety.upper()
    long_list, short_list = [], []
    contract = ""
    for r in rows:
        if not isinstance(r, dict):
            continue
        inst = str(r.get("INSTRUMENTID", "") or "")
        if var_u and var_u not in inst.upper():
            continue
        try:
            rank = int(float(r.get("RANK", 0) or 0))
        except (ValueError, TypeError):
            rank = len(long_list) + 1
        if rank <= 0:
            # SHFE 品种合计行（RANK=-1, INSTRUMENTID="rball"），跳过
            continue
        # 取首个真实合约名（跳过 aggregate 如 "rball"）
        if not contract:
            contract = inst
        long_member = str(r.get("PARTICIPANTABBR2", "") or "").strip()
        short_member = str(r.get("PARTICIPANTABBR3", "") or "").strip()
        try:
            long_lots = int(float(r.get("CJ2", 0) or 0))
        except (ValueError, TypeError):
            long_lots = 0
        try:
            short_lots = int(float(r.get("CJ3", 0) or 0))
        except (ValueError, TypeError):
            short_lots = 0
        if long_member and long_lots:
            long_list.append({"rank": rank, "member": long_member, "lots": long_lots})
        if short_member and short_lots:
            short_list.append({"rank": rank, "member": short_member, "lots": short_lots})
    if not long_list and not short_list:
        return {}
    lt = sum(x["lots"] for x in long_list)
    st = sum(x["lots"] for x in short_list)
    return {
        "long": long_list,
        "short": short_list,
        "net_position": lt - st,
        "contract": contract,
    }


def _parse_cffex_rank(text_or_bytes: Union[str, bytes], variety: str) -> dict:
    """解析中金所 ``{VAR}_1.csv``（GBK 编码，双行表头）。

    真实列（按位置）：
        [1]=合约, [2]=排名, [3]=会员简称(成交), [4]=成交量,
        [6]=会员简称(持买), [7]=持买单量,
        [9]=会员简称(持卖), [10]=持卖单量
    """
    if isinstance(text_or_bytes, bytes):
        text = text_or_bytes.decode("gbk", errors="replace")
    else:
        text = text_or_bytes
    # 按行拆，跳过前 2 行表头
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 3:
        return {}
    var_u = variety.upper()
    long_list, short_list = [], []
    contract = ""
    for ln in lines[2:]:  # skip header
        cols = [c.strip() for c in ln.split(",")]
        if len(cols) < 11:
            continue
        sym = cols[1]
        if var_u and var_u not in sym.upper():
            continue
        if not contract:
            contract = sym
        try:
            rank = int(float(cols[2]))
        except (ValueError, TypeError):
            rank = len(long_list) + 1
        long_member = cols[6].strip(" \"'")
        short_member = cols[9].strip(" \"'")
        def _to_int(v):
            try:
                return int(float(v.replace(",", "").strip(' "\'')))
            except (ValueError, TypeError):
                return 0
        long_lots = _to_int(cols[7])
        short_lots = _to_int(cols[10])
        if long_member and long_lots:
            long_list.append({"rank": rank, "member": long_member, "lots": long_lots})
        if short_member and short_lots:
            short_list.append({"rank": rank, "member": short_member, "lots": short_lots})
    if not long_list and not short_list:
        return {}
    lt = sum(x["lots"] for x in long_list)
    st = sum(x["lots"] for x in short_list)
    return {"long": long_list, "short": short_list, "net_position": lt - st, "contract": contract}


# ── CZCE 持仓排名（xlsx GET）───────────────────────────────────────────────

def _parse_czce_rank_xlsx(raw: bytes, variety: str) -> dict:
    """解析郑商所 FutureDataHolding.xlsx。

    xlsx 结构：标题行 → "品种：XXX"行 → 列标题行 → 数据行(20) → "合计"行 → 下一品种。
    直接定位 "品种：" 标记获取各品种块。
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c or "") for c in row])
    wb.close()
    if not all_rows:
        return {}

    var_u = variety.upper()
    long_list, short_list = [], []
    contract = ""

    for i, row in enumerate(all_rows):
        if not row or not row[0]:
            continue
        cell = row[0]
        # 找到 "品种：XXX" 行
        if "品种：" not in cell:
            continue
        if var_u and var_u not in cell.upper():
            continue
        # 提取合约名
        for token in re.findall(r"[A-Za-z0-9_]+", cell):
            contract = token
            break
        # 数据从该行 +2（跳过列标题行）开始，到下一个 "合计" 结束
        data_end = len(all_rows)
        for j in range(i + 1, len(all_rows)):
            if all_rows[j] and any("合计" in c for c in all_rows[j]):
                data_end = j
                break
        data_start = i + 2
        for r in all_rows[data_start:data_end]:
            if not r or not r[0] or "合计" in r[0]:
                break
            try:
                rank = int(float(str(r[0]).replace(",", "")))
            except (ValueError, TypeError):
                rank = len(long_list) + 1
            long_member = str(r[4] if len(r) > 4 else "").strip()
            short_member = str(r[7] if len(r) > 7 else "").strip()
            try:
                long_lots = int(float(str(r[5] if len(r) > 5 else 0).replace(",", "")))
            except (ValueError, TypeError):
                long_lots = 0
            try:
                short_lots = int(float(str(r[8] if len(r) > 8 else 0).replace(",", "")))
            except (ValueError, TypeError):
                short_lots = 0
            if long_member and long_lots:
                long_list.append({"rank": rank, "member": long_member, "lots": long_lots})
            if short_member and short_lots:
                short_list.append({"rank": rank, "member": short_member, "lots": short_lots})
        break  # found our variety, stop

    if not long_list and not short_list:
        return {}
    lt = sum(x["lots"] for x in long_list)
    st = sum(x["lots"] for x in short_list)
    return {"long": long_list, "short": short_list, "net_position": lt - st, "contract": contract}
    lt = sum(x["lots"] for x in long_list)
    st = sum(x["lots"] for x in short_list)
    return {"long": long_list, "short": short_list, "net_position": lt - st, "contract": contract}


# ── DCE 持仓排名（txt→excel GET，逐合约）────────────────────────────────────

def _parse_dce_rank_text(text: str, variety: str) -> dict:
    """解析大商所 txt 返回（从 HTML tables 降级解析）。

    列：名次, 会员简称, 成交量, 增减, 会员简称.1, 持买单量, 增减.1,
    会员简称.2, 持卖单量, 增减.2
    """
    return _parse_dce_rank_data(_parse_html(text), variety)


def _parse_dce_rank_xlsx(raw: bytes, variety: str) -> dict:
    """解析大商所 xlsx 返回（主路径）。"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        return {}
    # xlsx: header 4 行然后数据
    data_rows = []
    for r in rows:
        row = [str(c or "").strip() for c in r]
        if row and row[0].isdigit():
            data_rows.append(row)
    parsed_rows = []
    for r in data_rows:
        if len(r) >= 10:
            parsed_rows.append({
                "vol_party_name": r[1],
                "long_party_name": r[4],
                "long_open_interest": r[5],
                "short_party_name": r[7],
                "short_open_interest": r[8],
                "rank": r[0],
            })
    return _parse_dce_rank_data(parsed_rows, variety)


def _parse_dce_rank_data(rows: list[dict], variety: str) -> dict:
    """DCE 解析入口（共享给 txt/xlsx）。"""
    var_u = variety.upper()
    long_list, short_list = [], []
    contract = ""
    for i, r in enumerate(rows[:20]):
        try:
            rank = int(float(r.get("rank", 0)))
        except (ValueError, TypeError):
            rank = i + 1
        long_member = str(r.get("long_party_name", "") or "").strip()
        short_member = str(r.get("short_party_name", "") or "").strip()
        try:
            long_lots = int(float(str(r.get("long_open_interest", 0) or 0).replace(",", "")))
        except (ValueError, TypeError):
            long_lots = 0
        try:
            short_lots = int(float(str(r.get("short_open_interest", 0) or 0).replace(",", "")))
        except (ValueError, TypeError):
            short_lots = 0
        if long_member and long_lots:
            long_list.append({"rank": rank, "member": long_member, "lots": long_lots})
        if short_member and short_lots:
            short_list.append({"rank": rank, "member": short_member, "lots": short_lots})
    if not long_list and not short_list:
        return {}
    lt = sum(x["lots"] for x in long_list)
    st = sum(x["lots"] for x in short_list)
    return {"long": long_list, "short": short_list, "net_position": lt - st, "contract": contract}


# ── GFEX 持仓排名（JSON POST，逐合约 3 数据页合并）────────────────────────────

def _parse_gfex_rank(text: str, variety: str) -> dict:
    """解析广期所已合并 JSON（``_merge_gfex_pages`` 输出）。"""
    try:
        data = json.loads(text)
    except (ValueError, TypeError):
        return {}
    if not isinstance(data, dict):
        return {}
    rows = data.get("data") or data.get("o_cursor") or []
    if not isinstance(rows, list):
        return {}
    long_list, short_list = [], []
    for r in rows:
        if not isinstance(r, dict):
            continue
        try:
            rank = int(float(r.get("rank", 0) or 0))
        except (ValueError, TypeError):
            rank = len(long_list) + 1
        member = str(r.get("abbr", "") or "").strip()
        try:
            long_lots = int(float(r.get("long_open_interest", 0) or 0))
        except (ValueError, TypeError):
            long_lots = 0
        try:
            short_lots = int(float(r.get("short_open_interest", 0) or 0))
        except (ValueError, TypeError):
            short_lots = 0
        if member and long_lots:
            long_list.append({"rank": rank, "member": member, "lots": long_lots})
        if member and short_lots:
            short_list.append({"rank": rank, "member": member, "lots": short_lots})
    if not long_list and not short_list:
        return {}
    lt = sum(x["lots"] for x in long_list)
    st = sum(x["lots"] for x in short_list)
    return {"long": long_list, "short": short_list, "net_position": lt - st, "contract": ""}



def get_exchange_url(exchange: str, trade_date: str) -> Optional[str]:
    """根据交易所与交易日构造数据 URL；未知交易所返回 ``None``。"""
    ex = EXCHANGE_ENDPOINTS.get(exchange.upper())
    if ex is None:
        return None
    y = trade_date[:4]
    ym = trade_date[:6]
    day = trade_date[6:]
    ep = (
        ex["endpoint"]
        .replace("{year}", y)
        .replace("{year_month}", ym)
        .replace("{date}", trade_date)
        .replace("{day}", day)
    )
    return ex["base_url"] + ep


def get_warrant_url(exchange: str, trade_date: str) -> Optional[str]:
    """仓单日报专用URL；未知交易所返回 ``None``。"""
    ex = WARRANT_ENDPOINTS.get(exchange.upper())
    if ex is None:
        return None
    y = trade_date[:4]
    ym = trade_date[:6]
    day = trade_date[6:]
    ep = (
        ex["endpoint"]
        .replace("{year}", y)
        .replace("{year_month}", ym)
        .replace("{date}", trade_date)
        .replace("{day}", day)
    )
    return ex["base_url"] + ep


def fmt_of(exchange: str) -> Optional[str]:
    """返回交易所数据格式（json/csv/tsv/html）。"""
    ex = EXCHANGE_ENDPOINTS.get(exchange.upper())
    return ex["fmt"] if ex else None


def warrant_fmt_of(exchange: str) -> Optional[str]:
    """返回仓单日报数据格式（json/csv/tsv/html/xlsx）。"""
    ex = WARRANT_ENDPOINTS.get(exchange.upper())
    return ex["fmt"] if ex else None


async def fetch_exchange_page(
    url: str,
    *,
    transport: Optional[Callable[[str], Union[str, Awaitable[str]]]] = None,
    headers: Optional[dict] = None,
    as_bytes: bool = False,
) -> Union[str, bytes]:
    """抓取交易所页面文本（可注入 transport）。

    Args:
        url: 目标 URL。
        transport: 注入的抓取器（同步或异步，返回页面文本）；缺省使用 httpx。
        headers: 注入请求头（如 SHFE 需特定 UA）。
        as_bytes: True 时返回原始字节（xlsx 等二进制格式）；False 返回文本。

    Raises:
        Exception: 真实 httpx 请求失败时上抛（调用方自行降级）。
    """
    if transport is not None:
        out = transport(url)
        if hasattr(out, "__await__"):
            out = await out
        return out  # type: ignore[return-value]
    if as_bytes:
        return await _httpx_get_bytes(url, headers=headers)
    return await _httpx_get(url, headers=headers)


async def _httpx_get(url: str, headers: Optional[dict] = None) -> str:
    """真实 httpx GET 返回文本（抽离为模块级函数，便于测试桩替换）。"""
    import httpx

    async with httpx.AsyncClient(timeout=15, headers=headers or {},
                                 follow_redirects=True) as client:
        resp = await client.get(url)
        resp.raise_for_status()
        return resp.text


async def _httpx_get_bytes(url: str, headers: Optional[dict] = None) -> bytes:
    """httpx GET 返回原始字节（用于 xlsx 等二进制格式）。"""
    import httpx

    async with httpx.AsyncClient(timeout=15, headers=headers or {},
                                 follow_redirects=True) as client:
        resp = await client.get(url)
        resp.raise_for_status()
        return resp.content


async def _httpx_post(url: str, data: Optional[dict] = None, headers: Optional[dict] = None) -> str:
    """httpx POST 返回文本。"""
    import httpx

    async with httpx.AsyncClient(timeout=15, headers=headers or {},
                                 follow_redirects=True) as client:
        resp = await client.post(url, data=data or {})
        resp.raise_for_status()
        return resp.text


def parse_daily_rows(text: str, fmt: str) -> list[dict]:
    """将交易所日线文本解析为行字典列表（纯函数）。"""
    if not text:
        return []
    fmt = (fmt or "").lower()
    if fmt == "json":
        return _parse_json(text)
    if fmt == "csv":
        return _parse_csv(text, ",")
    if fmt == "tsv":
        return _parse_csv(text, "\t")
    if fmt == "html":
        return _parse_html(text)
    return []


def _parse_json(text: str) -> list[dict]:
    try:
        data = json.loads(text)
    except (ValueError, TypeError):
        return []
    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        # SHFE kx 格式常为 {"o_curinstrument": [...]} 等单一列表值
        lists = [v for v in data.values() if isinstance(v, list)]
        rows = lists[0] if lists else []
    else:
        return []
    return [r for r in rows if isinstance(r, dict)]


def _parse_csv(text: str, delimiter: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [dict(row) for row in reader]


def _parse_html(text: str) -> list[dict]:
    if not _HAS_BS4:
        return []
    soup = BeautifulSoup(text, "html.parser")
    table = soup.find("table")
    if table is None:
        return []
    rows = table.find_all("tr")
    if not rows:
        return []
    header_cells = rows[0].find_all(["th", "td"])
    headers = [c.get_text(strip=True) for c in header_cells]
    result: list[dict] = []
    for tr in rows[1:]:
        cells = tr.find_all("td")
        if len(cells) != len(headers):
            continue
        result.append({headers[i]: cells[i].get_text(strip=True) for i in range(len(headers))})
    return result


def today_yyyymmdd() -> str:
    """返回 ``YYYYMMDD`` 格式当前交易日。"""
    return datetime.now().strftime("%Y%m%d")
